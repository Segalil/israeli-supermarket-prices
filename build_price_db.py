#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Build an Excel/CSV price database from Israeli supermarket price-transparency data.

Downloads the official price files (Israel's "Food Law" price transparency) for the
*online* stores of the leading grocery chains, parses them, and writes a combined
Excel + CSV dataset — plus a promotions (PromoFull) CSV.

Run from a machine with **Israeli** internet egress where possible: several chains'
publishing sites (laibcatalog.co.il, shop.hazi-hinam.co.il, …) geo-block or WAF-block
foreign/cloud IPs, so GitHub Actions runners only reach part of the chains.

    pip install -r requirements.txt
    python build_price_db.py                    # 6 leading chains, online store
    python build_price_db.py --chains SHUFERSAL RAMI_LEVY
    python build_price_db.py --no-promos        # prices only
    python build_price_db.py --all-stores       # every branch (very large)
"""
import argparse
import glob
import os
import re
import shutil
import sys
import time
from datetime import datetime

from israeli_prices.parser import (
    parse_price_file,
    parse_promo_file,
    parse_store_file,
    looks_online,
)

# The leading chains with real online stores: scraper key -> Hebrew label.
# Smaller chains (e.g. TIV_TAAM, HAZI_HINAM) can be added via --chains.
LEADING_CHAINS = {
    "SHUFERSAL": "שופרסל",
    "RAMI_LEVY": "רמי לוי",
    "VICTORY_NEW_SOURCE": "ויקטורי",
    "YAYNO_BITAN_AND_CARREFOUR": "יינות ביתן / קרפור",
    "YOHANANOF": "יוחננוף",
    "OSHER_AD": "אושר עד",
}

DUMP_DIR = "dumps"
OUT_DIR = "out"
RETRIES = 2

COL_ORDER = ["chain", "chain_id", "store_id", "barcode", "item_name",
             "manufacturer", "country", "quantity", "unit_of_measure",
             "unit_qty", "price", "unit_price", "is_weighted",
             "qty_in_package", "allow_discount", "item_status",
             "price_update_date"]
COL_HEB = {
    "chain": "רשת", "chain_id": "מזהה רשת", "store_id": "מזהה חנות",
    "barcode": "ברקוד", "item_name": "שם מוצר", "manufacturer": "יצרן",
    "country": "ארץ ייצור", "quantity": "כמות", "unit_of_measure": "יחידת מידה",
    "unit_qty": "יחידת כמות", "price": "מחיר (₪)", "unit_price": "מחיר ליחידה (₪)",
    "is_weighted": "שקיל", "qty_in_package": "כמות באריזה",
    "allow_discount": "מותר הנחה", "item_status": "סטטוס", "price_update_date": "עודכן",
}

PROMO_COL_ORDER = ["chain", "chain_id", "store_id", "barcode", "promo_id",
                   "description", "end_date", "min_qty", "discounted_price",
                   "discount_rate", "club", "is_coupon", "reward_type"]
PROMO_COL_HEB = {
    "chain": "רשת", "chain_id": "מזהה רשת", "store_id": "מזהה חנות",
    "barcode": "ברקוד", "promo_id": "מזהה מבצע", "description": "תיאור מבצע",
    "end_date": "בתוקף עד", "min_qty": "כמות מינימום",
    "discounted_price": "מחיר מבצע (₪)", "discount_rate": "שיעור הנחה",
    "club": "מועדון", "is_coupon": "קופון", "reward_type": "סוג הטבה",
}


# --------------------------- scraping ---------------------------
def _scrape_targeted(chain_key, file_type, limit=None, store_id=None):
    """Scrape one chain/file-type, optionally filtered to a single store id.

    Uses the runner internals of il-supermarket-scraper (pinned in
    requirements.txt) because the public ScarpingTask API does not expose
    ``store_id``. Falls back to the public API (no store filter) if the
    internals move.
    """
    from il_supermarket_scarper import FileTypesFilters
    ftype = getattr(FileTypesFilters, file_type).name
    try:
        from il_supermarket_scarper.scrapper_runner import scrape_one_wrap
        from il_supermarket_scarper.utils.databases import (
            create_file_output_for_scraper,
            create_status_database_for_scraper,
        )
        kwargs = {
            "limit": limit,
            "files_types": [ftype],
            "single_pass": True,
            "file_output": create_file_output_for_scraper(chain_key, None),
            "status_database": create_status_database_for_scraper(chain_key, None),
        }
        if store_id:
            kwargs["store_id"] = int(store_id)
        scrape_one_wrap(chain_key, kwargs)
    except ImportError:
        from il_supermarket_scarper import ScarpingTask
        task = ScarpingTask(enabled_scrapers=[chain_key],
                            files_types=[ftype], multiprocessing=1)
        task.start(limit=limit)
        task.join()


def scrape(chain_key, file_type, limit=None, store_id=None):
    """_scrape_targeted with retries (transient blocks/timeouts are common)."""
    last = None
    for attempt in range(1 + RETRIES):
        try:
            _scrape_targeted(chain_key, file_type, limit=limit, store_id=store_id)
            return
        except Exception as exc:                      # noqa: BLE001
            last = exc
            if attempt < RETRIES:
                wait = 5 * (attempt + 1)
                print(f"    scrape retry in {wait}s ({exc})")
                time.sleep(wait)
    raise last


def chain_dump_folder(chain_key):
    """Locate the dump folder created for a chain."""
    try:
        from il_supermarket_scarper.utils import DumpFolderNames
        p = os.path.join(DUMP_DIR, DumpFolderNames[chain_key].value)
        if os.path.isdir(p):
            return p
    except Exception:
        pass
    subs = [os.path.join(DUMP_DIR, d) for d in os.listdir(DUMP_DIR)
            if os.path.isdir(os.path.join(DUMP_DIR, d)) and d != "status"]
    return max(subs, key=os.path.getmtime) if subs else None


def find_online_store(chain_key):
    """Download a stores file; return (online_store_id | None, stores_found_count)."""
    try:
        scrape(chain_key, "STORE_FILE", limit=1)
        folder = chain_dump_folder(chain_key)
        if not folder:
            return None, 0
        stores = []
        for f in glob.glob(os.path.join(folder, "*")):
            if "store" in os.path.basename(f).lower():
                stores.extend(parse_store_file(f))
        for sid, name in stores:
            if looks_online(name):
                return sid, len(stores)
        return None, len(stores)
    except Exception as exc:
        print(f"    (store file / online detection failed: {exc})")
        return None, 0


def _files_of_type(folder, kind):
    """Price/promo files in a dump folder ('price' excludes 'promo' and stores)."""
    out = []
    for f in glob.glob(os.path.join(folder or "", "*")):
        base = os.path.basename(f).lower()
        if "store" in base:
            continue
        if kind == "promo" and "promo" in base:
            out.append(f)
        elif kind == "price" and "price" in base and "promo" not in base:
            out.append(f)
    return out


def _file_store_matches(path, store_id):
    """True if a filename belongs to the given store id (dash/underscore segment)."""
    sid = store_id.lstrip("0") or "0"
    for seg in re.split(r"[-_.]", os.path.basename(path)):
        if seg.isdigit() and (seg.lstrip("0") or "0") == sid:
            return True
    return False


def _download_kind(chain_key, kind, online_id, all_stores):
    """Download price/promo files for a chain, preferring the online store.

    Returns (files_to_parse, note). Tries a store-targeted scrape first; if
    that yields nothing (some chains number files differently), falls back to
    an unfiltered small download and picks a representative file.
    """
    ftype = "PRICE_FULL_FILE" if kind == "price" else "PROMO_FULL_FILE"
    folder = chain_dump_folder(chain_key)

    if all_stores:
        scrape(chain_key, ftype)
        return _files_of_type(chain_dump_folder(chain_key), kind), "all stores"

    if online_id:
        try:
            scrape(chain_key, ftype, store_id=online_id)
        except Exception as exc:
            print(f"    {kind} targeted download error: {exc}")
        folder = chain_dump_folder(chain_key)
        files = [f for f in _files_of_type(folder, kind)
                 if _file_store_matches(f, online_id)]
        if files:
            # portals often expose weeks of history per store — parse only the
            # newest file (date is embedded in the name, so max() = newest)
            return [max(files)], f"online store {online_id}"
        print(f"    {kind}: store-targeted download empty – falling back.")

    try:
        scrape(chain_key, ftype, limit=2)
    except Exception as exc:
        print(f"    {kind} download error: {exc}")
    folder = chain_dump_folder(chain_key)
    files = _files_of_type(folder, kind)
    if online_id:
        matched = [f for f in files if _file_store_matches(f, online_id)]
        if matched:
            return [max(matched)], f"online store {online_id}"
    return ([max(files)], "representative branch") if files else ([], "no files")


def collect_chain(chain_key, label, all_stores=False, with_promos=True):
    """Return (price_rows, promo_rows) for one chain."""
    print(f"[{label}] starting…")
    online_id, stores_found = (None, 0)
    if not all_stores:
        online_id, stores_found = find_online_store(chain_key)
        print(f"    stores found: {stores_found} | online store_id: {online_id or 'not detected'}")

    price_files, note = _download_kind(chain_key, "price", online_id, all_stores)
    print(f"    price files: {len(price_files)} ({note})")
    rows = []
    for f in price_files:
        try:
            rows.extend(parse_price_file(f, label))
        except Exception as exc:
            print(f"    parse error on {os.path.basename(f)}: {exc}")
    print(f"    collected {len(rows):,} price rows.")

    promo_rows = []
    if with_promos and rows:
        # promos only matter for stores we actually price
        store_ids = {r.get("store_id", "") for r in rows}
        promo_files, pnote = _download_kind(chain_key, "promo", online_id, all_stores)
        promo_files = [f for f in promo_files
                       if all_stores or any(_file_store_matches(f, s) for s in store_ids if s)] \
            or promo_files[:1]
        for f in promo_files:
            try:
                promo_rows.extend(parse_promo_file(f, label))
            except Exception as exc:
                print(f"    promo parse error on {os.path.basename(f)}: {exc}")
        print(f"    collected {len(promo_rows):,} promo rows ({pnote}).")
    return rows, promo_rows


# --------------------------- output ---------------------------
def write_outputs(rows, stamp):
    import pandas as pd
    os.makedirs(OUT_DIR, exist_ok=True)
    df = pd.DataFrame(rows)
    for c in COL_ORDER:
        if c not in df.columns:
            df[c] = ""
    df = df[COL_ORDER]
    for c in ("price", "unit_price"):
        df[c] = pd.to_numeric(df[c], errors="coerce")

    csv_path = os.path.join(OUT_DIR, f"israeli_prices_{stamp}.csv")
    df.rename(columns=COL_HEB).to_csv(csv_path, index=False, encoding="utf-8-sig")

    xlsx_path = os.path.join(OUT_DIR, f"israeli_prices_{stamp}.xlsx")
    _write_xlsx(df, xlsx_path)
    return csv_path, xlsx_path, df


def write_promo_outputs(promo_rows, stamp):
    import pandas as pd
    os.makedirs(OUT_DIR, exist_ok=True)
    df = pd.DataFrame(promo_rows)
    for c in PROMO_COL_ORDER:
        if c not in df.columns:
            df[c] = ""
    df = df[PROMO_COL_ORDER]
    for c in ("club", "is_coupon"):
        df[c] = df[c].map(lambda v: 1 if v in (True, 1, "1") else 0)
    csv_path = os.path.join(OUT_DIR, f"israeli_promos_{stamp}.csv")
    df.rename(columns=PROMO_COL_HEB).to_csv(csv_path, index=False, encoding="utf-8-sig")
    return csv_path, df


def _write_xlsx(df, path):
    import pandas as pd
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    summary = (df.groupby("chain")
                 .agg(מוצרים=("barcode", "count"),
                      מחיר_חציוני=("price", "median"),
                      מחיר_ממוצע=("price", "mean"))
                 .reset_index().rename(columns={"chain": "רשת"}))
    with pd.ExcelWriter(path, engine="openpyxl") as xl:
        df.rename(columns=COL_HEB).to_excel(xl, sheet_name="מחירים", index=False)
        summary.to_excel(xl, sheet_name="סיכום", index=False)
        for ws in xl.book.worksheets:
            ws.sheet_view.rightToLeft = True
            fill = PatternFill("solid", fgColor="1F4E78")
            for cell in ws[1]:
                cell.font = Font(name="Arial", bold=True, color="FFFFFF")
                cell.fill = fill
                cell.alignment = Alignment(horizontal="center", vertical="center")
            ws.freeze_panes = "A2"
            ws.auto_filter.ref = ws.dimensions
            for col in ws.columns:
                width = max((len(str(c.value)) for c in col if c.value), default=8)
                ws.column_dimensions[get_column_letter(col[0].column)].width = min(width + 3, 45)
            for row in ws.iter_rows(min_row=2):
                for c in row:
                    if c.font is None or c.font.name != "Arial":
                        c.font = Font(name="Arial")


# --------------------------- main ---------------------------
def main():
    ap = argparse.ArgumentParser(description=__doc__)
    ap.add_argument("--chains", nargs="*", default=list(LEADING_CHAINS),
                    help="chain scraper keys (default: the 6 leading chains; "
                         "extras like TIV_TAAM / HAZI_HINAM can be added)")
    ap.add_argument("--all-stores", action="store_true",
                    help="every branch instead of just the online store (very large)")
    ap.add_argument("--no-promos", dest="promos", action="store_false",
                    help="skip PromoFull download/parsing")
    args = ap.parse_args()

    if os.path.isdir(DUMP_DIR):
        shutil.rmtree(DUMP_DIR, ignore_errors=True)

    stamp = datetime.now().strftime("%Y%m%d")
    all_rows, all_promos = [], []
    per_chain = {}
    for key in args.chains:
        label = LEADING_CHAINS.get(key, key)
        try:
            rows, promos = collect_chain(key, label, all_stores=args.all_stores,
                                         with_promos=args.promos)
        except Exception as exc:
            print(f"[{label}] error: {exc}")
            rows, promos = [], []
        per_chain[label] = len(rows)
        all_rows.extend(rows)
        all_promos.extend(promos)

    print("\n=== coverage ===")
    for label, n in per_chain.items():
        print(f"  [{'ok  ' if n else 'MISS'}] {label}: {n:,} rows")
    missing = [l for l, n in per_chain.items() if n == 0]
    if missing:
        print(f"  chains with no data: {', '.join(missing)}")
        print("  (note: some chains geo-block foreign/cloud IPs — run from Israeli "
              "egress for full coverage)")

    if not all_rows:
        print("No data collected. Check internet access and chain keys.")
        sys.exit(1)

    csv_path, xlsx_path, df = write_outputs(all_rows, stamp)
    print(f"\ntotal rows: {len(df):,}")
    print(f"CSV : {csv_path}")
    print(f"XLSX: {xlsx_path}")
    if args.promos:
        promo_csv, pdf = write_promo_outputs(all_promos, stamp)
        print(f"PROMOS: {promo_csv} ({len(pdf):,} rows)")


if __name__ == "__main__":
    main()
